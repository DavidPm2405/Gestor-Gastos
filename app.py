from flask import Flask, render_template, redirect, url_for, request, flash, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, Transaction
from datetime import datetime, date, timedelta
import csv
import io
import calendar
import secrets
import os

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "cambiar-en-produccion")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///gastos.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Flask-Mail (configurar con variables de entorno)
app.config["MAIL_SERVER"]   = "smtp.gmail.com"
app.config["MAIL_PORT"]     = 587
app.config["MAIL_USE_TLS"]  = True
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME", "")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD", "")

db.init_app(app)
mail = Mail(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Inicia sesión para continuar."

# ── Monedas ──────────────────────────────────────────────────────────────────

CURRENCIES = {
    "CLP": {"symbol": "$",   "code": "CLP", "name": "Peso Chileno",    "decimals": 0},
    "USD": {"symbol": "US$", "code": "USD", "name": "Dólar Americano", "decimals": 2},
    "EUR": {"symbol": "€",   "code": "EUR", "name": "Euro",            "decimals": 2},
    "ARS": {"symbol": "$",   "code": "ARS", "name": "Peso Argentino",  "decimals": 0},
    "MXN": {"symbol": "$",   "code": "MXN", "name": "Peso Mexicano",   "decimals": 2},
    "COP": {"symbol": "$",   "code": "COP", "name": "Peso Colombiano", "decimals": 0},
    "PEN": {"symbol": "S/.", "code": "PEN", "name": "Sol Peruano",     "decimals": 2},
    "BRL": {"symbol": "R$",  "code": "BRL", "name": "Real Brasileño",  "decimals": 2},
}

@app.template_filter("money")
def money_filter(amount, currency="CLP"):
    curr = CURRENCIES.get(currency, CURRENCIES["CLP"])
    if curr["decimals"] == 0:
        return f"{curr['symbol']}{int(amount):,}".replace(",", ".")
    return f"{curr['symbol']}{amount:,.{curr['decimals']}f}"

@app.context_processor
def inject_globals():
    currency = "CLP"
    if current_user.is_authenticated:
        currency = current_user.currency or "CLP"
    return {"user_currency": currency, "CURRENCIES": CURRENCIES}

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.route("/")
@login_required
def index():
    return redirect(url_for("dashboard"))


# ── Auth ─────────────────────────────────────────────────────────────────────

@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form["username"].strip()
        email    = request.form["email"].strip().lower()
        password = request.form["password"]
        if User.query.filter((User.username == username) | (User.email == email)).first():
            flash("El usuario o email ya existe.", "danger")
        elif len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "warning")
        else:
            user = User(username=username, email=email,
                        password_hash=generate_password_hash(password))
            db.session.add(user)
            db.session.commit()
            login_user(user)
            return redirect(url_for("dashboard"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user, remember=True)
            return redirect(request.args.get("next") or url_for("dashboard"))
        flash("Usuario o contraseña incorrectos.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Recuperar contraseña ──────────────────────────────────────────────────────

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        user  = User.query.filter_by(email=email).first()
        if user:
            token  = secrets.token_urlsafe(32)
            user.reset_token        = token
            user.reset_token_expiry = datetime.utcnow() + timedelta(hours=1)
            db.session.commit()
            reset_url = url_for("reset_password", token=token, _external=True)
            try:
                msg = Message("Recuperar contraseña — GastosApp",
                    sender=app.config["MAIL_USERNAME"],
                    recipients=[email])
                msg.body = (f"Hola {user.username},\n\n"
                            f"Haz clic en el siguiente enlace para restablecer tu contraseña:\n\n"
                            f"{reset_url}\n\n"
                            f"El enlace expira en 1 hora.\n\n"
                            f"Si no solicitaste esto, ignora este mensaje.")
                mail.send(msg)
                flash("Te enviamos un email con las instrucciones.", "success")
            except Exception:
                flash("No se pudo enviar el email. Contacta al administrador.", "danger")
        else:
            # No revelar si el email existe o no
            flash("Si ese email está registrado, recibirás instrucciones.", "info")
        return redirect(url_for("login"))
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    user = User.query.filter_by(reset_token=token).first()
    if not user or user.reset_token_expiry < datetime.utcnow():
        flash("El enlace es inválido o ha expirado.", "danger")
        return redirect(url_for("forgot_password"))
    if request.method == "POST":
        password = request.form["password"]
        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "warning")
        else:
            user.password_hash      = generate_password_hash(password)
            user.reset_token        = None
            user.reset_token_expiry = None
            db.session.commit()
            flash("Contraseña actualizada. Ya puedes iniciar sesión.", "success")
            return redirect(url_for("login"))
    return render_template("reset_password.html", token=token)


# ── Cambiar moneda ────────────────────────────────────────────────────────────

@app.route("/settings/currency", methods=["POST"])
@login_required
def change_currency():
    currency = request.form.get("currency", "CLP")
    if currency in CURRENCIES:
        current_user.currency = currency
        db.session.commit()
        flash(f"Moneda cambiada a {CURRENCIES[currency]['name']}.", "success")
    return redirect(request.referrer or url_for("dashboard"))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    today = date.today()
    year  = request.args.get("year",  today.year,  type=int)
    month = request.args.get("month", today.month, type=int)

    txs = Transaction.query.filter_by(user_id=current_user.id).all()

    total_income  = sum(t.amount for t in txs if t.type == "income")
    total_expense = sum(t.amount for t in txs if t.type == "expense")
    balance       = total_income - total_expense

    month_txs     = [t for t in txs if t.date.year == year and t.date.month == month]
    month_income  = sum(t.amount for t in month_txs if t.type == "income")
    month_expense = sum(t.amount for t in month_txs if t.type == "expense")

    monthly_income, monthly_expense = [], []
    for m in range(1, 13):
        monthly_income.append(sum(t.amount for t in txs if t.type == "income"  and t.date.year == year and t.date.month == m))
        monthly_expense.append(sum(t.amount for t in txs if t.type == "expense" and t.date.year == year and t.date.month == m))

    cat_data = {}
    for t in month_txs:
        if t.type == "expense":
            cat_data[t.category] = cat_data.get(t.category, 0) + t.amount

    recent     = sorted(txs, key=lambda t: t.date, reverse=True)[:5]
    month_name = calendar.month_name[month]
    years      = sorted({t.date.year for t in txs} | {today.year}, reverse=True)

    return render_template("dashboard.html",
        balance=balance, total_income=total_income, total_expense=total_expense,
        month_income=month_income, month_expense=month_expense,
        monthly_income=monthly_income, monthly_expense=monthly_expense,
        cat_labels=list(cat_data.keys()), cat_values=list(cat_data.values()),
        recent=recent, year=year, month=month, month_name=month_name,
        years=years, months=list(range(1, 13)), calendar=calendar)


# ── Transactions ──────────────────────────────────────────────────────────────

CATEGORIES = {
    "income":  ["Salario", "Freelance", "Inversiones", "Regalo", "Otro"],
    "expense": ["Alimentación", "Transporte", "Vivienda", "Salud",
                "Entretenimiento", "Ropa", "Educación", "Servicios", "Otro"],
}

@app.route("/transactions")
@login_required
def transactions():
    q           = Transaction.query.filter_by(user_id=current_user.id)
    type_filter = request.args.get("type", "")
    cat_filter  = request.args.get("category", "")
    if type_filter:
        q = q.filter_by(type=type_filter)
    if cat_filter:
        q = q.filter_by(category=cat_filter)
    txs      = q.order_by(Transaction.date.desc()).all()
    all_cats = CATEGORIES["income"] + CATEGORIES["expense"]
    return render_template("transactions.html", transactions=txs,
        type_filter=type_filter, cat_filter=cat_filter, all_cats=all_cats)


@app.route("/transactions/add", methods=["GET", "POST"])
@login_required
def add_transaction():
    if request.method == "POST":
        try:
            amount = float(request.form["amount"])
            if amount <= 0:
                raise ValueError
        except ValueError:
            flash("El monto debe ser un número positivo.", "warning")
            return redirect(url_for("add_transaction"))
        tx = Transaction(
            user_id=current_user.id,
            type=request.form["type"],
            category=request.form["category"],
            amount=amount,
            description=request.form.get("description", "").strip(),
            date=datetime.strptime(request.form["date"], "%Y-%m-%d").date(),
        )
        db.session.add(tx)
        db.session.commit()
        flash("Transacción agregada.", "success")
        return redirect(url_for("transactions"))
    return render_template("add_transaction.html",
        categories=CATEGORIES, today=date.today().isoformat())


@app.route("/transactions/delete/<int:tx_id>", methods=["POST"])
@login_required
def delete_transaction(tx_id):
    tx = db.get_or_404(Transaction, tx_id)
    if tx.user_id != current_user.id:
        flash("No autorizado.", "danger")
        return redirect(url_for("transactions"))
    db.session.delete(tx)
    db.session.commit()
    flash("Transacción eliminada.", "success")
    return redirect(url_for("transactions"))


# ── Export ────────────────────────────────────────────────────────────────────

@app.route("/export/csv")
@login_required
def export_csv():
    txs    = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Tipo", "Categoría", "Descripción", "Monto", "Moneda"])
    for t in txs:
        tipo = "Ingreso" if t.type == "income" else "Gasto"
        writer.writerow([t.date, tipo, t.category, t.description, t.amount, current_user.currency])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv", as_attachment=True, download_name="gastos.csv")


@app.route("/export/excel")
@login_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("Instala openpyxl: pip install openpyxl", "warning")
        return redirect(url_for("transactions"))

    txs = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Transacciones"

    headers     = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto", "Moneda"]
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(txs, 2):
        tipo = "Ingreso" if t.type == "income" else "Gasto"
        ws.append([str(t.date), tipo, t.category, t.description, t.amount, current_user.currency])
        ws.cell(row=row, column=5).font = Font(color="217346" if t.type == "income" else "C00000")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="gastos.xlsx")


with app.app_context():
    db.create_all()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
